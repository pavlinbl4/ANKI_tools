import re
from openpyxl import load_workbook
from pathlib import Path


def file_name_with_suffix(path_to_xlsx_file, suffix):
    return f'{Path(path_to_xlsx_file).parent}/{Path(path_to_xlsx_file).stem}_{suffix}{Path(path_to_xlsx_file).suffix}'


def remove_html_from_cells(path_to_xlsx_file):
    # Load the Excel file
    workbook = load_workbook(path_to_xlsx_file)
    sheet = workbook.active

    # Define a function to remove HTML tags
    def remove_html_tags(text):
        clean = re.compile('<.*?>')
        return re.sub(clean, '', text)

    # Iterate through each cell in columns A to I and remove HTML tags
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=8):
        for cell in row:
            if cell.value:
                cell.value = remove_html_tags(str(cell.value))
    # Save the modified workbook
    workbook.save(file_name_with_suffix(path_to_xlsx_file, 'cleared'))


if __name__ == '__main__':
    # /Users/evgeniy/Library/Mobile Documents/com~apple~CloudDocs/ANKI
    print(file_name_with_suffix((Path().home()/'Library/Mobile Documents/com~apple~CloudDocs/ANKI/All Decks.xlsx'), 'test_suffix'))

