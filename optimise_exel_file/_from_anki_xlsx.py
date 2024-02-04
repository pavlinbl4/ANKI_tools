import openpyxl
from optimise_exel_file import cleanhtml
from pathlib import Path


def read_data(xlsx_file):
    file_data = openpyxl.load_workbook(xlsx_file)
    return file_data


def get_sheet_names(xlsx_file_data):
    return xlsx_file_data.sheetnames


def find_specific_cell(current_sheet, columns_for_edit):
    for row in range(8, current_sheet.max_row + 1):
        for column in columns_for_edit:  # Here you can add or reduce the columns
            cell_name: str = "{}{}".format(column, row)
            if current_sheet[cell_name].value is not None:
                if "<" in current_sheet[cell_name].value:
                    # print("cell position {} has value {}".format(cell_name, current_sheet[cell_name].value))
                    clean_value = cleanhtml(current_sheet[cell_name].value)  # remove html tags
                    print(clean_value)
                    current_sheet[cell_name].value = clean_value


def clear_cells_from_html(input_xlsx_file, columns_for_edit):  # main function
    xlsx_file_data = read_data(input_xlsx_file)
    sheet_names = get_sheet_names(xlsx_file_data)
    current_sheet = xlsx_file_data[sheet_names[0]]
    find_specific_cell(current_sheet, columns_for_edit)
    xlsx_file_data.save(f'{Path().home()}/Documents/ANKI/free_from_tags.xlsx')


if __name__ == '__main__':
    anki_excel_file = f'{Path().home()}/Documents/ANKI/ANKI_EXCEL.xlsx'
    columns = "BCDEF"
    clear_cells_from_html(anki_excel_file, columns)
