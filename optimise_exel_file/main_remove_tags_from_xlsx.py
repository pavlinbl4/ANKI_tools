from pathlib import Path

import openpyxl
from remove_tags import cleanhtml
from tkinter import filedialog, messagebox
import os


def read_data(xlsx_file):
    file_data = openpyxl.load_workbook(xlsx_file)
    return file_data


def get_sheet_names(xlsx_file_data):
    return xlsx_file_data.sheetnames


def find_specific_cell(current_sheet):
    for row in range(8, current_sheet.max_row + 1):
        for column in "ABCDEGH":  # Here you can add or reduce the columns
            cell_name: str = "{}{}".format(column, row)
            if current_sheet[cell_name].value is not None:
                if "<" in current_sheet[cell_name].value:
                    # print("cell position {} has value {}".format(cell_name, current_sheet[cell_name].value))
                    clean_value = cleanhtml(current_sheet[cell_name].value).strip()
                    print(clean_value)
                    current_sheet[cell_name].value = clean_value


def iterate_cells(current_sheet):
    # Iterate through each cell in columns A to I and remove HTML tags
    for row in current_sheet.iter_rows(min_row=1, min_col=1, max_col=9):
        for cell in row:
            if cell.value:
                cell.value = "$$$$$$$"


def clear_from_tags(anki_excel_file):
    file_name = get_filename(anki_excel_file)
    xlsx_file_data = read_data(anki_excel_file)
    sheet_names = get_sheet_names(xlsx_file_data)
    current_sheet = xlsx_file_data[sheet_names[0]]
    find_specific_cell(current_sheet)

    # overwrite source file
    # xlsx_file_data.save(anki_excel_file)

    messagebox.showinfo(title="All done", message=f'file {file_name}_cleared.xlsx \ncreated')


def select_file():  # return path to the file
    return filedialog.askopenfilename(initialdir=(Path().home() / 'Library/Mobile Documents/com~apple~CloudDocs/ANKI'),
                                      defaultextension='xlsx')


def get_filename(file_path):
    return os.path.basename(file_path).split('.')[0]


if __name__ == '__main__':
    # clear_from_tags(select_file())
    clear_from_tags((Path().home() / 'Library/Mobile Documents/com~apple~CloudDocs/ANKI/All Decks_cleared.xlsx'))
